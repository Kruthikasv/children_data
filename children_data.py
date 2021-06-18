import glob
from pathlib import Path
import shutil
from distutils.dir_util import copy_tree
import os
import os.path
import csv
import re
import regex
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np

class user_metadata :
    def __init__(self,user,input_type):
        self.user = user
        self.input_type=input_type
    # matching different audio files done by a single user
    def matching_audio(self):
        user_name = []
        user_dict = dict()
        for i in self.user:
            for j in self.user:
                if Path(i).parts[10] == Path(j).parts[10]:
                    user_name.append(str(Path(j)))
            user_dict.update({str(Path(i).parts[10]): user_name})
        return(user_dict)
    def extraction(self,user_dict):             #extracting all data of one particular user
        data_list = []
        final_data = []
        with open('/home/kruthika/Desktop/children_data/firestore-backup.json', 'r') as f:
            data_full = f.readlines()[0]
            for i in user_dict.keys():
                start_index = data_full.index(str(i))
                end_index = data_full[start_index:].index("}}}}") + start_index
                data_list = (data_full[start_index - 1:end_index + 1].split(","))
                final_data.append(data_list)
        return(data_list,final_data)
    def search_attributes(self,data_list,final_data):
        # creating a list which contains attributes of anonymous users
        metadata_list = []
        metadata_list.append(data_list[0].rstrip("\""))
        # clear metadata_list before use to flush out duplicate data
        metadata_list.clear()
        # finding the attribute values
        for i in final_data:
            metadata_list.append(i[0].split(":")[0])
            for j in i:
                subj_id = re.findall("\"subjectId\":\"[a-z|A-Z|0-9]*\"", j)
                if len(subj_id) != 0:
                    metadata_list.append(subj_id[0].split(":"))
            for j in i:
                name_object = re.findall("\"name\":\"[a-z|A-Z|0-9|\.|\_|\-|\:|\'| ]*\"", j)
                if len(name_object) != 0:
                    metadata_list.append(name_object[0].split(":"))
            k = 0
            for j in i:
                email_object = re.findall("\"email\":\"[^@]+@[^@]+\.[^@]+\"", j)
                if len(email_object) != 0:
                    k = 1
                    metadata_list.append(email_object[0].split(":"))
            if k != 1:
                metadata_list.append("[\"NA\"]")
            for j in i:
                phone_object = re.findall("\"phoneNo\":\"[0-9]*\"", j)
                if len(phone_object) != 0:
                    metadata_list.append(phone_object[0].split(":"))
            for j in i:
                yob_object = re.findall("\"yob\":\"[0-9]*\"", j)
                if len(yob_object) != 0:
                    metadata_list.append(yob_object[0].split(":"))
                else:
                    yob_object = re.findall("\"yob\":[0-9]*", j)
                    if len(yob_object) != 0:
                        metadata_list.append(yob_object[0].split(":"))
            for j in i:
                mob_object = re.findall("\"mob\":\"[a-z|A-Z]*\"", j)
                if len(mob_object) != 0:
                    metadata_list.append(mob_object[0].split(":"))
            for j in i:
                school_object = re.findall("\"school\":\"[a-z,|A-Z,|0-9|\.|\_|\-|\:|\'|,|\&|\(|\)| ,| ]*", j)
                if len(school_object) != 0:
                    metadata_list.append(school_object[0].split(":"))
            for j in i:
                class_object = re.findall("\"croom\":\"[a-z,|A-Z,|0-9|\.|\_|\-|\:|\'|,|\&|\(|\)| ,| ]*", j)
                if len(class_object) != 0:
                    metadata_list.append(class_object[0].split(":"))
            for j in i:
                origin_object = re.findall("\"origin\":\"[a-z|A-Z| ]*\"", j)
                if len(origin_object) != 0:
                    metadata_list.append(origin_object[0].split(":"))
            for j in i:
                mt_object = re.findall("\"mt\":\"[a-z|A-Z]*\"", j)
                if len(mt_object) != 0:
                    metadata_list.append(mt_object[0].split(":"))
            for j in i:
                gender_object = re.findall("\"gender\":\"[a-z|A-Z]*\"", j)
                if len(gender_object) != 0:
                    metadata_list.append(gender_object[0].split(":"))
            lang = []
            all_lang = ""
            for j in i:
                rlang_object = re.findall("\"language\":\"[a-z|A-Z]*\"", j)
                if len(rlang_object) != 0:
                    lang.append(rlang_object[0].split(":"))
            for index in range(len(lang)):
                all_lang = all_lang + " " + lang[index][1].strip('\"')
            metadata_list.append(all_lang)
        return(metadata_list)

    def save_data(self,metadata_list):             # saving required data in xl for anonymous users
        wb = Workbook()
        ws1 = wb.create_sheet("anonymous")
        c = 1
        r = 2
        for i in metadata_list:
            j = i[1:]
            if (type(j) == str):
                m = j.rstrip('\"')
                if len(m) == 28 or len(m) == 20:
                    r += 1
                    c = 1
            ws1.cell(row=r, column=c).value = str(j).strip("['\"\"']")
            c = c + 1

        wb.save('/home/kruthika/Desktop/children_data/agender/udata.xlsx')
    def save_data_named(self,metadata_list):           # saving required data in xl for named users
        wb2 = load_workbook('/home/kruthika/Desktop/children_data/agender/udata.xlsx')
        ws1 = wb2.create_sheet("named")
        c = 1
        r = 2
        for i in metadata_list:
            j = i[1:]
            if (type(j) == str):
                m = j.rstrip('\"')
                if len(m) == 28 or len(m) == 20:
                    r += 1
                    c = 1
            ws1.cell(row=r, column=c).value = str(j).strip("['\"\"']")
            c = c + 1
        wb2.save('/home/kruthika/Desktop/children_data/agender/udata.xlsx')

    def user_data(self,final_data):
        userdata_list = []
        userdata_list.clear()
        for i in final_data:
            userdata_list.append(i[0].split(":")[0])
            for j in i:
                name_object = re.findall("\"name\":\"[a-z,A-Z0-9\.|\_|\-:.| ]*\"", j)
                if len(name_object) != 0:
                    userdata_list.append(name_object[0].split(":"))

            for j in i:
                age_object = re.findall("\"age\":[0-9]*", j)
                if len(age_object) != 0:
                    userdata_list.append(age_object[0].split(":"))

            for j in i:
                subj_id = re.findall("\"subjectId\":\"[a-z|A-Z|0-9]*\"", j)
                if len(subj_id) != 0:
                    userdata_list.append(subj_id[0].split(":"))
        j = 0
        length = len(userdata_list)
        id_name_age = []
        userdata = []
        while j < length:
            if type(userdata_list[j]) == str:
                id_name_age.clear()
                id_name_age.append(userdata_list[j].strip("\""))
                id_name_age.append(userdata_list[j + 1][1].strip("\""))
                id_name_age.append(userdata_list[j + 2][1])
                id_name_age.append(userdata_list[j + 3][1].strip("\""))
                userdata.append(id_name_age[0:4])    # appending provider_id, name, age and subject_id
            j += 4
        return(userdata)

    def age_classificaton(self,userdata):       #separating users based on age (below age 12 and above)
        below_12 = []
        above_12 = []
        for i in userdata:
            if int(i[2]) <= 12:
                below_12.append(i)
            else:
                above_12.append(i)
        return(above_12,below_12)

    def language(self):
        lang_names = ['assamese', 'bengali', 'bodo', 'dogri', 'english', 'gujarati', 'hindi', 'kannada', 'kashmiri',
                      'konkani', 'maithili', 'malayalam', 'manipuri', 'marathi', 'nepali', 'odia', 'punjabi',
                      'sanskrit', 'santali', 'sindhi', 'tamil', 'telugu', 'urdu']

        # fetching all data for each language
        sentences = dict()
        with open('/home/kruthika/Desktop/children_data/firestore-backup.json', 'r') as f:
            languages = f.readlines()[0]
            for i in lang_names:
                start_index = languages.index(str(i))
                end_index = languages[start_index:].index("}}}}") + start_index
                lang_list = (languages[start_index - 1:end_index + 1].split(","))
                sentences.update({i: str(lang_list)})  # display language and sentences in dictionary format
        return(sentences)

    def count_recordings(self,userdata):
        wav_path = []
        ids = []
        count_list = []
        recording_count = {}
        count = 0
        for i in self.user:
            wav_path.append(Path(i).parts[-1].split("_"))
        for i in wav_path:
            if i[2] not in ids:
                ids.append(i[2])

        for i in ids:
            for j in wav_path:
                if i == j[2]:
                    count += 1
            count_list.append([i, count])
            count = 0
        for i in count_list:
            for j in userdata:
                if i[0] == j[0]:
                    i.append(int(j[2]))
        for i in count_list:
            recording_count.update({i[0]: i[1]})
        return (recording_count,wav_path)

    def make_directories(self,above_12,recording_count,wav_path,sentences):     #creating directories for all users above the age 12
        lang_with_sent = ['bengali', 'english', 'hindi', 'kannada', 'marathi', 'tamil', 'telugu']
        lang_img = ['assamese', 'bodo', 'dogri', 'gujarati', 'kashmiri', 'konkani', 'maithili', 'malayalam', 'manipuri',
                    'nepali', 'odia', 'punjabi', 'sanskrit', 'santali', 'sindhi', 'urdu']
        for i in above_12:
            #get_lang = []
            id_index = 0
            directory = i[3] + "_" + i[1]
            parent_dir = "/home/kruthika/Desktop/children_data/agender/"+self.input_type
            count = recording_count.get(str(i[3]))
            for j in wav_path:
                if i[3] == j[2]:
                    if count > 0:
                        get_lang = sentences.get(str(j[3]))  # matching the lang of user to lang in the metadata and getting all sentences
                        id_obj = re.findall("\"id\":[0-9]*", get_lang)
                        rec_lang = str(j[3])
                        s = '"id"' + ':' + j[4]
                        if s in id_obj:
                            id_index = id_obj.index(s)
                        if str(j[3]) in lang_with_sent:
                            path = os.path.join(parent_dir, directory, "sentences", rec_lang)
                            try:
                                os.makedirs(path, exist_ok=True)
                                # print("Directory '%s' created successfully" %directory)
                            except OSError as error:
                                print("Directory '%s' can not be created")
                            if s in id_obj:
                                id_index = id_obj.index(s)
                            sent_obj = regex.findall("\"sentence\":\"[\w'-|.|!| ]*", get_lang)
                            if (len(id_obj) != 0 and len(sent_obj) != 0):
                                sentence = sent_obj[id_index]
                                count -= 1
                                filename = "sentence" + "_" + j[3]
                                sf = open(path + "/" + filename + ".txt", "a")
                                sf.write(id_obj[id_index] + ", " + sentence + "\n")
                            else:
                                continue
                        elif str(j[3]) in lang_img:
                            image_id = []
                            path = "/home/kruthika/Desktop/children_data/Stimuli/Stimuli/adults"
                            dir_list = os.listdir(path)
                            for i in dir_list:
                                image_id.append(i.rstrip(".jpg"))
                            rec_lang = str(j[3])
                            path = os.path.join(parent_dir, directory, "images", rec_lang)
                            try:
                                os.makedirs(path, exist_ok=True)
                                # print("Directory '%s' created successfully" %directory)
                            except OSError as error:
                                print("Directory '%s' can not be created")
                            for k in image_id:
                                if k == j[4]:
                                    image = j[4] + ".jpg"
                                    images_path = "/home/kruthika/Desktop/children_data/Stimuli/Stimuli/adults/" + image
                                    shutil.copy2(images_path, path)

    def make_directories2(self,below_12,recording_count,wav_path):           #creating directories for all users below the age 12
        image_id = []
        path = "/home/kruthika/Desktop/children_data/Stimuli/Stimuli/children"
        dir_list = os.listdir(path)
        for i in dir_list:
            image_id.append(i.rstrip(".jpg"))
        for i in below_12:                 # making directories for all anon users below 12 and storing images
            directory = i[3] + "_" + i[1]
            parent_dir = "/home/kruthika/Desktop/children_data/agender/"+self.input_type
            count = recording_count.get(str(i[3]))
            for j in wav_path:
                if i[3] == j[2]:
                    if count > 0:
                        rec_lang = str(j[3])
                        path = os.path.join(parent_dir, directory, "images", rec_lang)
                        try:
                            os.makedirs(path, exist_ok=True)
                            # print("Directory '%s' created successfully" %directory)
                        except OSError as error:
                            print("Directory '%s' can not be created")
                        for k in image_id:
                            if k == j[4]:
                                image = j[4] + ".jpg"
                                images_path = "/home/kruthika/Desktop/children_data/Stimuli/Stimuli/children/" + image
                                shutil.copy2(images_path, path)

    def audio_folder(self):                 # getting the audio folders and storing them in respective anonymous user folders
        src_path = "/home/kruthika/Desktop/ALL_AUDIO"
        dest_path = "/home/kruthika/Desktop/children_data/agender/"+self.input_type
        src_list = os.listdir(src_path)
        dest_list = os.listdir(dest_path)
        dest_id = []
        for i in dest_list:
            dest_id.append(i.split("_")[0])
        for i in dest_list:
            _index = i.index("_")
            for j in src_list:
                if i[0:_index] == j:
                    audio_path = src_path + "/" + j
                    final_path = dest_path + "/" + i + "/audio"
                    copy_tree(audio_path, final_path)

    def update_workbook(self):
        dir_list = os.listdir("/home/kruthika/Desktop/children_data/agender/"+self.input_type)
        ws1 = load_workbook('/home/kruthika/Desktop/children_data/agender/udata.xlsx')
        sheet = ws1[self.input_type]
        row = sheet.max_row
        for i in range(1, row + 1):
            for j in dir_list:
                if sheet.cell(row=i, column=2).value == j.split("_")[0]:
                    sheet.cell(row=i, column=14).value = j
        ws1.save('/home/kruthika/Desktop/children_data/agender/udata.xlsx')

def delete_copies():
    dir = "/home/kruthika/Desktop/children_data/agender"
    for files in os.listdir(dir):
        path = os.path.join(dir, files)
        try:
            shutil.rmtree(path)
        except OSError:
            os.remove(path)
def main() :
    # accessing the paths of each audio file
    file_list = []
    for wave_file in glob.glob("/home/kruthika/Desktop/children_data/audios/**/*.wav", recursive=True):
        file_list.append(wave_file)
    df = pd.DataFrame(np.column_stack([file_list]))
    df.to_csv("paths.csv", index=False)

    # opening and reading the paths from csv file
    with open("paths.csv", "r") as f:
        reader = csv.reader(f, delimiter="\n")
        data = list(reader)
    # segregating the paths into anonymous and named
    anonymous = []
    named = []
    for i in data[1:]:
        if Path(i[0]).parts[-4] == "Anonymous":
            anonymous.append(Path(i[0]))
        else:
            named.append(Path(i[0]))
    delete_copies()
    exec(anonymous,"anonymous")
    exec(named,"named")

def exec(user,string):
    user_obj = user_metadata(user,string)
    output1=user_obj.matching_audio()
    output2,output3=user_obj.extraction(output1)
    output4=user_obj.search_attributes(output2,output3)
    if string=="anonymous":
        user_obj.save_data(output4)
    else:
        user_obj.save_data_named(output4)
    output_userdata=user_obj.user_data(output3)
    output_sentences=user_obj.language()
    output_above12,output_below_12=user_obj.age_classificaton(output_userdata)
    output_recording_count,output_wav_path=user_obj.count_recordings(output_userdata)
    user_obj.make_directories(output_above12,output_recording_count,output_wav_path,output_sentences)
    user_obj.make_directories2(output_below_12,output_recording_count,output_wav_path)
    user_obj.audio_folder()
    user_obj.update_workbook()

if __name__ =="__main__" :
    main()